import streamlit as st
import pandas as pd
import os
import platform
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font

# Inicializar estado
if "cantidades" not in st.session_state:
    st.session_state.cantidades = [None] * 7
if "observaciones_dia" not in st.session_state:
    st.session_state.observaciones_dia = [""] * 7
if "producto" not in st.session_state:
    st.session_state.producto = ""

# Guardar datos
def guardar_datos(df_nuevo):
    os.makedirs("data", exist_ok=True)
    archivo = "data/datos.csv"
    df_nuevo = df_nuevo[["fecha", "tipo", "producto", "cantidad", "observacion", "registrado_por"]]
    if os.path.exists(archivo):
        df_existente = pd.read_csv(archivo, sep=";")
        df_completo = pd.concat([df_existente, df_nuevo], ignore_index=True)
    else:
        df_completo = df_nuevo
    df_completo.to_csv(archivo, index=False, sep=";")

# Cargar datos
def cargar_datos():
    archivo = "data/datos.csv"
    if os.path.exists(archivo):
        return pd.read_csv(archivo, sep=";")
    return pd.DataFrame(columns=["fecha", "tipo", "producto", "cantidad", "observacion", "registrado_por"])

# Exportar Excel
def exportar_excel(df, inicio_semana):
    os.makedirs("data", exist_ok=True)
    nombre_archivo = f"reporte_{inicio_semana.strftime('%Y%m%d')}.xlsx"
    ruta = os.path.join("data", nombre_archivo)
    df = df[df["tipo"] == "produccion"]

    try:
        with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
            # Orden de columnas: A=fecha, B=producto, C=cantidad, D=registrado_por, E=observacion
            df_ordenado = df[["fecha", "producto", "cantidad", "registrado_por", "observacion"]]
            df_ordenado.to_excel(writer, sheet_name="Registro", index=False)

        wb = openpyxl.load_workbook(ruta)
        ws = wb["Registro"]
        ws.insert_rows(1, amount=5)
        ws["A1"] = "K-sa de K-mpo 8000"
        ws["A2"] = f"Semana: {inicio_semana.strftime('%d/%m/%Y')}"
        ws["A3"] = f"Producto: {df['producto'].iloc[0]}"
        ws["A5"] = "Producción semanal"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A5"].font = Font(bold=True, size=12)
        ws[f"A{ws.max_row + 3}"] = "Firma responsable:"
        wb.save(ruta)

    except Exception as e:
        st.error(f"❌ Error al generar Excel: {e}")
        return

    if platform.system() == "Windows":
        try:
            os.startfile(ruta)
        except Exception as e:
            st.warning(f"⚠️ No se pudo abrir automáticamente: {e}")
    else:
        st.info("ℹ️ Archivo generado. Abrilo manualmente desde la carpeta 'data'.")

# Interfaz principal
st.title("REGISTROS K-sa de K-mpo 8000")

# Fecha y producto
col1, col2 = st.columns([1, 2])
with col1:
    inicio_semana = st.date_input("Inicio de semana (lunes)", value=date.today())
with col2:
    opciones = ["Huevo", "Miel de Abeja", "Lechuga", "Perejil", "Tomate", "Otro"]
    seleccion = st.selectbox("Producto", opciones)
    if seleccion == "Otro":
        producto = st.text_input("Escribí el nombre del producto", value=st.session_state.producto)
    else:
        producto = seleccion
    st.session_state.producto = producto

# Registrado por
nombres = ["Mabel Enciso", "Richart Gonzalez", "Carmen Martinez", "Amalio Enciso", "Otro"]
seleccion_nombre = st.selectbox("Registrado por", nombres)
if seleccion_nombre == "Otro":
    registrador = st.text_input("Escribí tu nombre")
else:
    registrador = seleccion_nombre

# Cantidades y observaciones diarias
st.subheader("📊 Producción y observaciones diarias")
fechas = [inicio_semana + timedelta(days=i) for i in range(7)]
dias = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
cantidades = []
observaciones_dia = []
cols = st.columns(7)
for i, col in enumerate(cols):
    with col:
        st.markdown(f"**{dias[i]}**")
        st.markdown(f"{fechas[i].strftime('%d/%m')}")
        cantidad = st.number_input(
            label="Cantidad",
            min_value=0.0,
            format="%.4f",
            key=f"c_{i}",
            value=st.session_state.cantidades[i],
            placeholder="Cantidad"
        )
        observacion_dia = st.text_input(
            label="Obs.",
            value=st.session_state.observaciones_dia[i],
            key=f"obs_{i}",
            placeholder="Comentario"
        )
        cantidades.append(cantidad)
        observaciones_dia.append(observacion_dia)
        st.session_state.cantidades[i] = cantidad
        st.session_state.observaciones_dia[i] = observacion_dia

# Guardar
if st.button("Guardar"):
    registros = []
    for i in range(7):
        if cantidades[i] is not None and cantidades[i] > 0:
            registros.append({
                "fecha": fechas[i].strftime("%Y-%m-%d"),
                "tipo": "produccion",
                "producto": producto,
                "cantidad": cantidades[i],
                "observacion": observaciones_dia[i],
                "registrado_por": registrador
            })
    if registros:
        guardar_datos(pd.DataFrame(registros))
        st.success("✅ Datos guardados correctamente")
        st.session_state.cantidades = [None] * 7
        st.session_state.observaciones_dia = [""] * 7
        st.session_state.producto = ""
        st.rerun()
    else:
        st.warning("⚠️ No se registró ninguna cantidad.")

# Exportar
if st.button("📥 Exportar reporte semanal"):
    df = cargar_datos()
    exportar_excel(df, inicio_semana)
    st.success("✅ Excel generado con hoja única y encabezado de producto")